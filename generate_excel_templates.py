"""
اسکریپت ایجاد فایل‌های قالب اکسل و داده‌های نمونه
Generate Excel Templates and Sample Data

این اسکریپت فایل‌های اکسل قالب و داده نمونه برای تمام جداول مدل‌ها ایجاد می‌کند
This script creates Excel template files and sample data for all model tables
"""
import os
import pandas as pd
from datetime import datetime, timedelta
import random
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def create_directories():
    """ایجاد پوشه‌های مورد نیاز / Create required directories"""
    os.makedirs('excel_templates', exist_ok=True)
    os.makedirs('excel_sample_data', exist_ok=True)
    print("✓ Directories created: excel_templates/ and excel_sample_data/")


def create_excel_with_headers(filename, columns_mapping):
    """
    ایجاد فایل اکسل با هدرهای دو زبانه
    Create Excel file with bilingual headers
    
    Args:
        filename: نام فایل / File name
        columns_mapping: دیکشنری {English: Persian} / Dictionary of column mappings
    """
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    ws.append(list(columns_mapping.keys()))  # English headers
    ws.append(list(columns_mapping.values()))  # Persian headers
    
    # Style the header rows
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for row in range(1, 3):
        for col in range(1, len(columns_mapping) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(filename)
    print(f"  ✓ Created template: {filename}")


def generate_transaction_template():
    """ایجاد قالب جدول Transactions"""
    columns = {
        'Id': 'شناسه',
        'DocumentDate': 'تاریخ سند',
        'DocumentNumber': 'شماره سند',
        'DocumentDescription': 'شرح سند',
        'AccountCode': 'کد حساب',
        'TotalCode': 'کد کل',
        'SubsidiaryCode': 'کد معین',
        'Detail1Code': 'کد تفصیل 1',
        'Detail2Code': 'کد تفصیل 2',
        'Detail3Code': 'کد تفصیل 3',
        'Debit': 'بدهکار',
        'Credit': 'بستانکار',
        'Description': 'شرح',
        'CounterPartyDescription': 'شرح طرف حساب',
        'IsDeleted': 'حذف شده',
        'CreationTime': 'زمان ایجاد',
        # New fields added
        'CheckNumber': 'شماره چک',
        'CheckStatus': 'وضعیت چک',
        'AccountNumber': 'شماره حساب بانکی',
        'Payee': 'دریافت‌کننده',
        'TransactionID': 'شناسه تراکنش',
        'TransactionDate': 'تاریخ تراکنش',
        'TransactionType': 'نوع تراکنش',
        'ReferenceNumber': 'شماره مرجع',
        'EmployeeID': 'شناسه کارمند',
        'PayrollAmount': 'مبلغ حقوق',
        'VendorID': 'شناسه فروشنده',
        'VendorName': 'نام فروشنده',
        'CustomerID': 'شناسه مشتری',
        'OriginalAmount': 'مبلغ اصلی',
        'DiscountAmount': 'مبلغ تخفیف',
        'ItemID': 'شناسه کالا',
        'Quantity': 'مقدار',
        'BeginningInventory': 'موجودی اول دوره',
        'EndingInventory': 'موجودی پایان دوره',
        'EntryType': 'نوع ثبت',
        'EntryTime': 'زمان ثبت',
        'EnteredBy': 'ثبت‌کننده'
    }
    
    create_excel_with_headers('excel_templates/Transactions_Template.xlsx', columns)


def generate_check_payables_template():
    """ایجاد قالب جدول CheckPayables"""
    columns = {
        'Id': 'شناسه',
        'DocumentPaymentNumber': 'شماره برگه پرداخت',
        'DocumentPaymentDate': 'تاریخ برگه پرداخت',
        'CheckNumber': 'شماره چک',
        'CheckAmount': 'مبلغ چک',
        'CheckDate': 'تاریخ چک',
        'PayeeCode': 'کد دریافت‌کننده',
        'PayeeName': 'نام دریافت‌کننده',
        'IsDeleted': 'حذف شده',
        'CreationTime': 'زمان ایجاد'
    }
    
    create_excel_with_headers('excel_templates/CheckPayables_Template.xlsx', columns)


def generate_check_receivables_template():
    """ایجاد قالب جدول CheckReceivables"""
    columns = {
        'Id': 'شناسه',
        'DocumentReceiptNumber': 'شماره برگه دریافت',
        'DocumentReceiptDate': 'تاریخ برگه دریافت',
        'CheckNumber': 'شماره چک',
        'CheckAmount': 'مبلغ چک',
        'CheckDate': 'تاریخ چک',
        'DrawerCode': 'کد واگذارکننده',
        'DrawerName': 'نام واگذارکننده',
        'IsDeleted': 'حذف شده',
        'CreationTime': 'زمان ایجاد'
    }
    
    create_excel_with_headers('excel_templates/CheckReceivables_Template.xlsx', columns)


def generate_payroll_template():
    """ایجاد قالب جدول PayrollTransactions"""
    columns = {
        'Id': 'شناسه',
        'VoucherNumber': 'شماره سند',
        'VoucherDate': 'تاریخ سند',
        'Month': 'ماه',
        'EmployeeCode': 'کد کارمند',
        'EmployeeFullName': 'نام کارمند',
        'WorkedDays': 'روز کارکرد',
        'MissionDays': 'روز مأموریت',
        'OvertimeHours': 'ساعت اضافه‌کار',
        'BaseSalary': 'حقوق پایه',
        'OvertimePay': 'اضافه‌کاری',
        'InsuranceDeduction': 'بیمه',
        'TaxDeduction': 'مالیات',
        'NetPayment': 'خالص پرداختی',
        'IsDeleted': 'حذف شده',
        'CreationTime': 'زمان ایجاد'
    }
    
    create_excel_with_headers('excel_templates/PayrollTransactions_Template.xlsx', columns)


def generate_inventory_issues_template():
    """ایجاد قالب جدول InventoryIssues"""
    columns = {
        'Id': 'شناسه',
        'IssueNumber': 'شماره حواله',
        'IssueDate': 'تاریخ حواله',
        'ItemCode': 'کد کالا',
        'ItemName': 'نام کالا',
        'Quantity': 'مقدار',
        'UnitPrice': 'نرخ',
        'Amount': 'مبلغ',
        'CostCenterCode': 'کد مرکز هزینه',
        'CostCenterName': 'نام مرکز هزینه',
        'IsDeleted': 'حذف شده',
        'CreationTime': 'زمان ایجاد'
    }
    
    create_excel_with_headers('excel_templates/InventoryIssues_Template.xlsx', columns)


def generate_sales_transactions_template():
    """ایجاد قالب جدول SalesTransactions"""
    columns = {
        'Id': 'شناسه',
        'InvoiceNumber': 'شماره فاکتور',
        'InvoiceDate': 'تاریخ فاکتور',
        'CustomerCode': 'کد مشتری',
        'CustomerName': 'نام مشتری',
        'ItemCode': 'کد کالا',
        'ItemName': 'نام کالا',
        'Quantity': 'مقدار',
        'UnitPrice': 'نرخ',
        'Amount': 'مبلغ',
        'IsDeleted': 'حذف شده',
        'CreationTime': 'زمان ایجاد'
    }
    
    create_excel_with_headers('excel_templates/SalesTransactions_Template.xlsx', columns)


def generate_all_templates():
    """ایجاد تمام فایل‌های قالب"""
    print("\n📋 Creating Excel Templates...")
    print("=" * 60)
    
    generate_transaction_template()
    generate_check_payables_template()
    generate_check_receivables_template()
    generate_payroll_template()
    generate_inventory_issues_template()
    generate_sales_transactions_template()
    
    print("=" * 60)
    print("✅ All templates created successfully!\n")


def generate_transaction_sample_data():
    """ایجاد داده‌های نمونه برای Transactions"""
    num_records = 500
    
    # تاریخ شروع
    start_date = datetime(2024, 1, 1)
    
    # لیست کدهای حساب
    account_codes = ['1101', '1102', '1201', '2101', '2201', '3101', '4101', '5101', '6101']
    
    # لیست نام فروشندگان
    vendors = [
        ('V001', 'شرکت تأمین‌کننده الف'),
        ('V002', 'شرکت تأمین‌کننده ب'),
        ('V003', 'شرکت تأمین‌کننده ج'),
    ]
    
    # لیست مشتریان
    customers = [
        ('C001', 'C002', 'C003', 'C004', 'C005')
    ]
    
    # لیست کارمندان
    employees = ['E001', 'E002', 'E003', 'E004', 'E005', 'E006', 'E007', 'E008']
    
    # لیست اقلام
    items = ['ITEM001', 'ITEM002', 'ITEM003', 'ITEM004', 'ITEM005']
    
    data = []
    
    for i in range(num_records):
        # تاریخ تصادفی
        random_days = random.randint(0, 365)
        trans_date = start_date + timedelta(days=random_days)
        doc_date = trans_date
        
        # نوع تراکنش
        trans_types = ['Purchase', 'Sale', 'Payment', 'Receipt', 'Payroll', 'Manual']
        trans_type = random.choice(trans_types)
        
        # مبلغ
        amount = random.randint(100000, 50000000)
        
        # تصمیم بدهکار/بستانکار
        debit = amount if random.random() < 0.5 else 0
        credit = 0 if debit > 0 else amount
        
        row = {
            'Id': i + 1,
            'DocumentDate': doc_date.strftime('%Y-%m-%d'),
            'DocumentNumber': 1000 + i,
            'DocumentDescription': f'سند شماره {1000 + i}',
            'AccountCode': random.choice(account_codes),
            'TotalCode': random.choice(account_codes)[:2],
            'SubsidiaryCode': random.choice(account_codes),
            'Detail1Code': '',
            'Detail2Code': '',
            'Detail3Code': '',
            'Debit': debit,
            'Credit': credit,
            'Description': f'شرح تراکنش {i + 1}',
            'CounterPartyDescription': 'طرف حساب',
            'IsDeleted': 'False',
            'CreationTime': trans_date.strftime('%Y-%m-%d %H:%M:%S'),
        }
        
        # اضافه کردن فیلدهای جدید بر اساس نوع تراکنش
        if trans_type == 'Payroll' or random.random() < 0.2:
            row['EmployeeID'] = random.choice(employees)
            row['PayrollAmount'] = random.randint(10000000, 50000000)
            row['TransactionType'] = 'Payroll'
        elif trans_type == 'Purchase':
            vendor = random.choice(vendors)
            row['VendorID'] = vendor[0]
            row['VendorName'] = vendor[1]
            row['TransactionType'] = 'Purchase'
            row['ItemID'] = random.choice(items)
            row['Quantity'] = random.randint(1, 100)
        elif trans_type == 'Sale':
            row['CustomerID'] = random.choice(customers)
            row['TransactionType'] = 'Sale'
            row['ItemID'] = random.choice(items)
            row['Quantity'] = random.randint(1, 50)
            if random.random() < 0.3:  # 30% با تخفیف
                row['OriginalAmount'] = amount * 1.2
                row['DiscountAmount'] = amount * 0.2
        else:
            row['TransactionType'] = trans_type
        
        # فیلدهای چک برای تراکنش‌های پرداخت
        if trans_type in ['Payment', 'Receipt'] or random.random() < 0.15:
            row['CheckNumber'] = f'CHK{random.randint(1000000, 9999999)}'
            row['CheckStatus'] = random.choice(['Issued', 'Outstanding', 'Pending', 'Cleared'])
            row['AccountNumber'] = f'{random.randint(1000, 9999)}-{random.randint(100000, 999999)}'
            row['Payee'] = random.choice([v[1] for v in vendors])
        
        # فیلدهای دیگر
        row['TransactionID'] = f'TRX{i + 1:06d}'
        row['TransactionDate'] = trans_date.strftime('%Y-%m-%d')
        row['ReferenceNumber'] = f'REF{random.randint(1000, 9999)}'
        
        # فیلدهای ثبت دستی
        if trans_type == 'Manual' or random.random() < 0.1:
            row['EntryType'] = 'Manual'
            entry_hour = random.randint(0, 23)
            entry_time = trans_date.replace(hour=entry_hour, minute=random.randint(0, 59))
            row['EntryTime'] = entry_time.strftime('%Y-%m-%d %H:%M:%S')
            row['EnteredBy'] = random.choice(['user1', 'user2', 'user3', 'admin'])
        else:
            row['EntryType'] = 'Automatic'
        
        # فیلدهای موجودی
        if row.get('ItemID'):
            row['BeginningInventory'] = random.randint(100, 1000)
            row['EndingInventory'] = random.randint(50, 900)
        
        data.append(row)
    
    # ایجاد DataFrame
    df = pd.DataFrame(data)
    
    # ذخیره فایل
    df.to_excel('excel_sample_data/Transactions_SampleData.xlsx', index=False)
    print(f"  ✓ Created sample data: Transactions_SampleData.xlsx ({len(df)} records)")


def generate_check_payables_sample_data():
    """ایجاد داده‌های نمونه برای CheckPayables"""
    num_records = 50
    start_date = datetime(2024, 1, 1)
    
    data = []
    for i in range(num_records):
        random_days = random.randint(0, 365)
        check_date = start_date + timedelta(days=random_days)
        payment_date = check_date - timedelta(days=random.randint(1, 30))
        
        data.append({
            'Id': i + 1,
            'DocumentPaymentNumber': f'PAY{i + 1:05d}',
            'DocumentPaymentDate': payment_date.strftime('%Y-%m-%d'),
            'CheckNumber': f'{random.randint(1000000, 9999999)}',
            'CheckAmount': random.randint(1000000, 100000000),
            'CheckDate': check_date.strftime('%Y-%m-%d'),
            'PayeeCode': f'P{random.randint(1000, 9999)}',
            'PayeeName': f'دریافت‌کننده {i + 1}',
            'IsDeleted': 'False',
            'CreationTime': payment_date.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    df = pd.DataFrame(data)
    df.to_excel('excel_sample_data/CheckPayables_SampleData.xlsx', index=False)
    print(f"  ✓ Created sample data: CheckPayables_SampleData.xlsx ({len(df)} records)")


def generate_payroll_sample_data():
    """ایجاد داده‌های نمونه برای PayrollTransactions"""
    num_records = 100
    start_date = datetime(2024, 1, 1)
    
    employees = [
        ('E001', 'علی احمدی'),
        ('E002', 'محمد رضایی'),
        ('E003', 'فاطمه محمدی'),
        ('E004', 'زهرا کریمی'),
        ('E005', 'حسین صالحی'),
    ]
    
    data = []
    for i in range(num_records):
        emp = random.choice(employees)
        month_offset = i // len(employees)
        voucher_date = start_date + timedelta(days=30 * month_offset)
        
        base_salary = random.randint(15000000, 50000000)
        overtime_hours = random.randint(0, 50)
        overtime_pay = overtime_hours * random.randint(100000, 300000)
        insurance = base_salary * 0.07
        tax = (base_salary + overtime_pay) * 0.10
        
        data.append({
            'Id': i + 1,
            'VoucherNumber': f'PAY{voucher_date.year}{voucher_date.month:02d}{i + 1:03d}',
            'VoucherDate': voucher_date.strftime('%Y-%m-%d'),
            'Month': f'{voucher_date.year}-{voucher_date.month:02d}',
            'EmployeeCode': emp[0],
            'EmployeeFullName': emp[1],
            'WorkedDays': random.randint(20, 30),
            'MissionDays': random.randint(0, 5),
            'OvertimeHours': overtime_hours,
            'BaseSalary': base_salary,
            'OvertimePay': overtime_pay,
            'InsuranceDeduction': insurance,
            'TaxDeduction': tax,
            'NetPayment': base_salary + overtime_pay - insurance - tax,
            'IsDeleted': 'False',
            'CreationTime': voucher_date.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    df = pd.DataFrame(data)
    df.to_excel('excel_sample_data/PayrollTransactions_SampleData.xlsx', index=False)
    print(f"  ✓ Created sample data: PayrollTransactions_SampleData.xlsx ({len(df)} records)")


def generate_inventory_sample_data():
    """ایجاد داده‌های نمونه برای InventoryIssues"""
    num_records = 80
    start_date = datetime(2024, 1, 1)
    
    items = [
        ('ITEM001', 'کالا الف'),
        ('ITEM002', 'کالا ب'),
        ('ITEM003', 'کالا ج'),
        ('ITEM004', 'کالا د'),
    ]
    
    cost_centers = [
        ('CC001', 'مرکز هزینه تولید'),
        ('CC002', 'مرکز هزینه اداری'),
        ('CC003', 'مرکز هزینه فروش'),
    ]
    
    data = []
    for i in range(num_records):
        random_days = random.randint(0, 365)
        issue_date = start_date + timedelta(days=random_days)
        item = random.choice(items)
        cc = random.choice(cost_centers)
        qty = random.randint(1, 100)
        unit_price = random.randint(10000, 500000)
        
        data.append({
            'Id': i + 1,
            'IssueNumber': f'ISS{i + 1:05d}',
            'IssueDate': issue_date.strftime('%Y-%m-%d'),
            'ItemCode': item[0],
            'ItemName': item[1],
            'Quantity': qty,
            'UnitPrice': unit_price,
            'Amount': qty * unit_price,
            'CostCenterCode': cc[0],
            'CostCenterName': cc[1],
            'IsDeleted': 'False',
            'CreationTime': issue_date.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    df = pd.DataFrame(data)
    df.to_excel('excel_sample_data/InventoryIssues_SampleData.xlsx', index=False)
    print(f"  ✓ Created sample data: InventoryIssues_SampleData.xlsx ({len(df)} records)")


def generate_sales_sample_data():
    """ایجاد داده‌های نمونه برای SalesTransactions"""
    num_records = 120
    start_date = datetime(2024, 1, 1)
    
    customers = [
        ('C001', 'مشتری الف'),
        ('C002', 'مشتری ب'),
        ('C003', 'مشتری ج'),
        ('C004', 'مشتری د'),
    ]
    
    items = [
        ('ITEM001', 'محصول A'),
        ('ITEM002', 'محصول B'),
        ('ITEM003', 'محصول C'),
    ]
    
    data = []
    for i in range(num_records):
        random_days = random.randint(0, 365)
        invoice_date = start_date + timedelta(days=random_days)
        customer = random.choice(customers)
        item = random.choice(items)
        qty = random.randint(1, 50)
        unit_price = random.randint(50000, 2000000)
        
        data.append({
            'Id': i + 1,
            'InvoiceNumber': f'INV{i + 1:05d}',
            'InvoiceDate': invoice_date.strftime('%Y-%m-%d'),
            'CustomerCode': customer[0],
            'CustomerName': customer[1],
            'ItemCode': item[0],
            'ItemName': item[1],
            'Quantity': qty,
            'UnitPrice': unit_price,
            'Amount': qty * unit_price,
            'IsDeleted': 'False',
            'CreationTime': invoice_date.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    df = pd.DataFrame(data)
    df.to_excel('excel_sample_data/SalesTransactions_SampleData.xlsx', index=False)
    print(f"  ✓ Created sample data: SalesTransactions_SampleData.xlsx ({len(df)} records)")


def generate_all_sample_data():
    """ایجاد تمام داده‌های نمونه"""
    print("\n📊 Creating Sample Data Files...")
    print("=" * 60)
    
    generate_transaction_sample_data()
    generate_check_payables_sample_data()
    generate_payroll_sample_data()
    generate_inventory_sample_data()
    generate_sales_sample_data()
    
    print("=" * 60)
    print("✅ All sample data files created successfully!\n")


def main():
    """تابع اصلی / Main function"""
    print("=" * 60)
    print("Excel Template and Sample Data Generator")
    print("ایجاد فایل‌های قالب و داده نمونه اکسل")
    print("=" * 60)
    
    create_directories()
    generate_all_templates()
    generate_all_sample_data()
    
    print("\n" + "=" * 60)
    print("📁 Files created in:")
    print("   - excel_templates/     (Template files with headers)")
    print("   - excel_sample_data/   (Sample data files)")
    print("=" * 60)
    print("\n✅ All done! You can now use these files for testing.")
    print("✅ تمام فایل‌ها ایجاد شدند! اکنون می‌توانید از آنها استفاده کنید.")


if __name__ == "__main__":
    main()
